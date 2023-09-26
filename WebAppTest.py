from flask import Flask, render_template, request, redirect, url_for
import openpyxl
from openpyxl.styles import NamedStyle, Font
import pandas as pd
from datetime import datetime
import os

app = Flask(__name__)

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        # Get the uploaded files and output folder
        cash_book_audit_file = request.files["cash_book_audit_file"]
        site_listing_file = request.files["site_listing_file"]
        output_folder = "output"  # Change this to your desired output folder path

        if cash_book_audit_file and site_listing_file:
            # Save the uploaded files to the server
            cash_book_audit_file.save(os.path.join(output_folder, "cash_book_audit.xlsx"))
            site_listing_file.save(os.path.join(output_folder, "site_listing.xlsx"))

            # Execute your code with the uploaded files
            execute_code(
                os.path.join(output_folder, "cash_book_audit.xlsx"),
                os.path.join(output_folder, "site_listing.xlsx"),
                output_folder,
            )

    return render_template("index.html")

# Function to execute the code (same as in your original script)

# Function to execute the code
def execute_code(cash_book_audit_file, site_listing_file, output_folder):
    if cash_book_audit_file and site_listing_file and output_folder:
        # Load the Safe Count File
        SafeCounts = pd.read_excel(cash_book_audit_file)

        # Delete unnecessary columns
        columns_to_drop = ['VenueID', 'PostingType', 'CarriedForward', 'CarriedForward1', 'PostingValueAbs', 'NewBalance',
                           'ClerkID', 'Clerk_Name', 'MediaDescription', 'PostingReference', 'Direction']
        SafeCounts.drop(columns=columns_to_drop, inplace=True)

        # Rename Columns
        SafeCounts.rename(columns={'V_Description': 'Site', 'DepositDate': 'Date', 'PostingValue': 'Variance'},
                          inplace=True)

        # Correct the Date
        SafeCounts['Date'] = pd.to_datetime(SafeCounts['Date'], dayfirst=True)
        SafeCounts.sort_values(by='Date', inplace=True)

        # Remove Duplicate Counts in the same day
        SafeCounts.drop_duplicates(subset=['Date', 'Site'], inplace=True)

        # Create a new Excel sheet
        SafeCounts.to_excel(os.path.join(output_folder, 'Safe Count Reminders.xlsx'), sheet_name='CountTracker',
                            index=False)

        # Import the list of sites
        SiteListing = pd.read_excel(site_listing_file)

        # Create a sheet for the Aging Report
        SafeCounts = openpyxl.load_workbook(os.path.join(output_folder, 'Safe Count Reminders.xlsx'))
        Aging = SafeCounts.create_sheet('AgingReport')
        Aging['A1'] = 'Sites'

        # Write the Site Listing into Aging Report Sheet
        for row_num, row in enumerate(SiteListing['Site'], start=2):
            Aging.cell(row=row_num, column=1, value=row)

        # Save the modified Excel File
        SafeCounts.save(os.path.join(output_folder, 'Safe Count Reminders.xlsx'))

        # Get the days since the last count
        sheet = SafeCounts.active
        today = datetime.today()
        date_column = 'B'
        sheet['D1'] = 'Days Since Today'

        for row, cell in enumerate(sheet[date_column], start=1):
            if isinstance(cell.value, datetime):
                days_difference = (cell.value - today).days
                sheet[f'D{row}'] = days_difference

        SafeCounts.save(os.path.join(output_folder, 'Safe Count Reminders.xlsx'))

        # Bring Days since the last count into the aging sheet
        CountTracker = SafeCounts['CountTracker']
        AgingReport = SafeCounts['AgingReport']
        site_to_number = {}

        for row in CountTracker.iter_rows(min_row=2, values_only=True):
            site, number = row[0], row[3]
            site_to_number[site] = number

        AgingReport['B1'] = 'DaysSinceLast'  # Header for the new column
        general_format = NamedStyle(name="general_format")
        general_format.number_format = "General"

        for row_num, site in enumerate(AgingReport.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
            site = site[0]
            if site in site_to_number:
                number = site_to_number[site]
                AgingReport.cell(row=row_num, column=2, value=number).style = general_format

        SafeCounts.save(os.path.join(output_folder, 'Safe Count Reminders.xlsx'))

        # Fix up the formatting
        CountTrackerSheet = SafeCounts['CountTracker']
        AgingReportSheet = SafeCounts['AgingReport']
        HeadingFonts = Font(name='Calibri', bold=True, size=14)
        HeadingRow = 1

        for cell in CountTrackerSheet[HeadingRow]:
            cell.font = HeadingFonts

        for cell in AgingReportSheet[HeadingRow]:
            cell.font = HeadingFonts

        columns_to_adjust = ['A', 'B', 'C', 'D']

        for column in columns_to_adjust:
            max_length = 0

            for cell in CountTrackerSheet[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

            adjusted_width = (max_length + 2)
            CountTrackerSheet.column_dimensions[column].width = adjusted_width

        columns_to_adjust = ['A', 'B']

        for column in columns_to_adjust:
            max_length = 0

            for cell in AgingReportSheet[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

            adjusted_width = (max_length + 2)
            AgingReportSheet.column_dimensions[column].width = adjusted_width

        CountTrackerSheet.sheet_state = 'hidden'

        SafeCounts.save(os.path.join(output_folder, 'Safe Count Reminders.xlsx'))

        # Add a button to email the site
        SiteListing = openpyxl.load_workbook(site_listing_file)
        SiteListingSheet = SiteListing.active
        email_addresses = {}

        for row in SiteListingSheet.iter_rows(min_row=2, values_only=True):
            site_name, to_email, cc_email = row
            email_addresses[site_name] = {'to': to_email, 'cc': cc_email}

        AgingReportSheet = SafeCounts.active

        for row_number, row in enumerate(AgingReportSheet.iter_rows(min_row=2, max_row=AgingReportSheet.max_row, values_only=True), start=2):
            site_name = row[0]
            email_info = email_addresses.get(site_name)

            if email_info:
                to_email = email_info['to']
                cc_email = email_info['cc']
                mailto_link = f'=HYPERLINK("mailto:{to_email}?cc={cc_email}&subject={site_name}%20- Safe Count Reminder&body=kkkkk", "Click to Send Email")'
                AgingReportSheet.cell(row=row_number, column=3, value=mailto_link)
                AgingReportSheet.cell(row=row_number, column=3).style = "Hyperlink"

        SafeCounts.save(os.path.join(output_folder, 'Safe Count Reminders.xlsx'))
        SiteListing.close()
        SafeCounts.close()

# Function to select the 'Cash Book Audit' file
def select_cash_book_audit_file():
    cash_book_audit_file = filedialog.askopenfilename(title="Select Cash Book Audit Excel file", filetypes=[("Excel files", "*.xlsx")])
    if cash_book_audit_file:
        select_site_listing_file(cash_book_audit_file)

# Function to select the 'SiteListing' file and output folder
def select_site_listing_file(cash_book_audit_file):
    site_listing_file = filedialog.askopenfilename(title="Select SiteListing Excel file", filetypes=[("Excel files", "*.xlsx")])
    if site_listing_file:
        output_folder = filedialog.askdirectory(title="Select Output Folder")
        execute_code(cash_book_audit_file, site_listing_file, output_folder)

# Function to display the Help dialog
def show_help_dialog():
    help_window = tk.Toplevel(root)
    help_window.title("About")

    help_info = """
    Author: Matthew Capparelli
    Email: Matthew.Capparelli.EXT@Sodexo.com
    If you need to change any emails, just edit the site listing folder
    """

    help_label = tk.Label(help_window, text=help_info, padx=20, pady=20)
    help_label.pack()

# Create the main GUI window
root = tk.Tk()
root.title("Safe Count Reminders")

# Set the window size
root.geometry("400x200")  

# Create a label for instructions
instruction_label = tk.Label(root, text="1. Select Cash Book Audit \n2. Select SiteListing \n3. Select Output Folder \nAfter this, it will appear in your designated folder")
instruction_label.pack(pady=10)

# Create buttons for selecting files
cash_book_audit_button = tk.Button(root, text="Start!", command=select_cash_book_audit_file)
cash_book_audit_button.pack()

# Create a button to show the Help dialog
help_button = tk.Button(root, text="Help", command=show_help_dialog)
help_button.pack()

# Start the GUI main loop
root.mainloop()


if __name__ == "__main__":
    app.run(debug=True)
